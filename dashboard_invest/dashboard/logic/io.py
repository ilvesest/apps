# native imports
import re, datetime, os, traceback
from functools import wraps
from typing import Callable
from types import NoneType

# 3rd party imports
import requests
from bs4 import BeautifulSoup as BS
import numpy as np
import openpyxl

# pandas 
import pandas as pd
pd.options.mode.chained_assignment = None

### CONSTANTS ###
# none type exlusion list
NONE_LIKE_LIST = ["", " ", np.nan, 'nan', 'NA', None, 'none', 'None', NoneType]

### CLASSES ###
class DDF(pd.DataFrame):
    """Custom DF class which values are dictionaries of type {'value':value, 'color':color,...}"""
    
    EXCLUSIONS = NONE_LIKE_LIST
    
    def __init__(self, data=None, index=None, columns=None, dtype=None, copy=False):
        super().__init__(data=data, index=index, columns=columns, dtype=dtype, copy=copy)  
    
    @property
    def v(self):
        def extract_value(value_dict):
            return value_dict.get('value')
        return self.applymap(extract_value)

    def prop(self, property: str):
        def extract_property(value_dict):
            return value_dict.get(property)
        return self.applymap(extract_property)
    
    def getDdfDict(self, references_dict: dict[str, tuple[str,str,str]]) -> dict[str]:
        """Return dictionary of subset DDF-s based on reference dict.
        
        Args:

            references_dict (dict): Dictionary of df_name: tuple('method', 'string' , 'direction', int('col_idx1)).
                method has 2 options ['contains' and 'equals'], 'direction' has 3 options 
                ['one', 'down', 'up'] where one means only one line needs to be parsed, 'up' and 'down'
                respectively correspond to the parsing direction from the reference point.
            
        Returns:
            dict: Dictionary of {'df_name': df}"""
        
        
        def get_reference_row_col_idx(df: pd.DataFrame, pattern: str, method: str='contains') -> tuple[int, int]:
            """Find reference row and column numeric index values.

            Args:
                pattern (str): Character sequence or regular expression.
                method (str, optional): Finding reference via pattern within the text ('contains')
                    or equalling the value exactly ('equals'). Defaults to 'contains'.

            Raises:
                ValueError: Raises error if other than ['contains', 'equals'] is specified for the method.

            Returns:
                tuple[int, int]: Tuple of ['row_i', 'col_i']
            """
            
            
            # validate that method is correctly entered
            if method not in ['contains', 'equals']:
                raise ValueError(f"{method} can take only values: 'contains' or 'equals'!")
            
            if method == 'contains':
                index, column = np.where(df.apply(lambda x: x.str.contains(pattern) == True))
                return index[0], column[0]
            
            if method == 'equals':
                index, column = np.where(df == pattern)
                return index[0], column[0]
    
        def get_DDF(ddf: DDF, 
            row_idx1: int=None, 
            row_idx2:int=None, 
            col_idx1: int=None, 
            col_idx2: int=None,
            col_0: bool=True, 
            direction='down') -> DDF:
            """Slice DF till the first occuring empty row in given direction.

            Args:
                col_0 (bool, optional): If col_idx1 is actually first column. Defaults to True.
                direction (str, optional): Slice upwards or downwards from given row index. Defaults to 'down'.

            Raises:
                ValueError: If no row indices are specified.

            Returns:
                pd.DataFrame
            """
            
            # assert that at least one of the row indices is specified
            if row_idx1 is None and row_idx2 is None:
                raise ValueError(f"Both row indices can't equal {None}!") 
            
            # if column index is not the first column then None
            col_idx1 = col_idx1 if col_0 is True else None
            
            # find missing row index
            if direction == 'down':
                nan_mask = ddf.v.loc[row_idx1:,col_idx1:col_idx2].isna().all(axis='columns')
                row_idx2 = None if nan_mask.sum() == 0 else nan_mask.idxmax()

            if direction == 'up':
                nan_mask = ddf.v.loc[:row_idx2,col_idx1:col_idx2].isna().all(axis='columns')
                row_idx1 = None if nan_mask.sum() == 0 else nan_mask[::-1].idxmax() + 1
                row_idx2 += 1
                
            return DDF(ddf.loc[row_idx1:row_idx2,col_idx1:col_idx2])
        
        
        # set all cols str type, NaN -> 'nan'
        df = self.v.astype('O')
        
        # {ddf name: ddf} dictionary
        ddfs = {}
        for k,v in references_dict.items():
            
            # find reference position (row index & col index)
            row_i, col_i = get_reference_row_col_idx(df, pattern=v[1], method=v[0])
            
            # shift ref position if specified
            col_i = v[3] if len(v) == 4 else col_i
            
            if v[2] == 'one':
                ddfs[k] = DDF(self.loc[row_i,col_i:])
            if v[2] == 'down':
                ddfs[k] = get_DDF(self, row_idx1=row_i, col_idx1=col_i, direction=v[2])
            if v[2] == 'up':
                ddfs[k] = get_DDF(self, row_idx2=row_i, col_idx1=col_i, direction=v[2])

        # strip all NaN cols and rows
        ddfs_clean = {}
        for k,ddf_ in ddfs.items():
            
            # drop any NaN-s in the row
            if isinstance(ddf_.v, pd.Series):
                df_ = ddf_.v.dropna(how='any', axis='rows')
                ddfs_clean[k] = DDF(ddf_.loc[df_.index,0])
            # drop 
            else:
                df_ = (ddf_.v 
                    .dropna(how='all', axis='columns')
                    .dropna(how='all', axis='rows')
                )
                ddfs_clean[k] = DDF(ddf_.loc[df_.index, df_.columns])
        
        return ddfs_clean
    
    def addButton(self, keys: list=[],
                  col_names: list[str]=None,
                  button_name: str="Details",
                  button_class: str="btn btn-secondary btn-sm",
                  data_bs_html: str="true",
                  data_bs_toggle: str="popover",
                  data_bs_trigger: str="focus",
                  button_style: str="--bs-btn-font-size: .85rem;"):
        """Add button html to the value-dict."""
        def add_button(value_dict):
            
            # select only specified keys and values that are not none-like
            style_args = [f"{key}:{val}" for key,val in value_dict.items() \
                if key in keys and val not in DDF.EXCLUSIONS]
            
            # concat style args into one string
            style_str = ";".join(style_args) if len(style_args) > 0 else ""
            data_value = value_dict['value'] 
                      
            html_string = " ".join(
                (
                f'<button type="button" class="{button_class}"', 
                f"""data-bs-content='<div style="{style_str}">{data_value}</div>'""",
                f'data-bs-html="{data_bs_html}" data-bs-toggle={data_bs_toggle} data-bs-trigger="{data_bs_trigger}"', 
                f'style="{button_style}">{button_name}</button>'
                )
            )
            return {'value':html_string} if data_value not in DDF.EXCLUSIONS else {'value':""}
        
        # apply to specified columns
        self[col_names] = DDF(self[col_names]).applymap(add_button)
        
        # return entire DDF
        return DDF(self)
        
    def replaceValues(self, df:pd.DataFrame, keys: list=None):
        """Replace values in DDF based on values in DF. Return DDF with DF dimensions."""
        df_idx, df_cols = df.index, df.columns
        for i in df_idx:
            for c in df_cols:
                self.loc[i,c]['value'] = df.loc[i,c]
        return DDF(self.loc[df_idx,df_cols])

    def replaceProperties(self, subset: pd.IndexSlice, property_dict: dict):
        """Replace values in DDF based on property_dict keys and values."""
        
        def replace_values(value_dict) -> dict:
            """Returns dictionary with updated properties."""
            new_value_dict = {k:v for k,v in value_dict.items() if k not in property_dict.keys()}
            return new_value_dict | property_dict
        
        if isinstance(self.loc[subset], pd.Series):
            self.loc[subset] = self.loc[subset].apply(replace_values)
            return DDF(self)
        else:
            self.loc[subset] = self.loc[subset].applymap(replace_values)
            return DDF(self)
               

    def setHeader(self, header_idx: int=None, col_names: list[str]=None):
        """Returns DDF with specified header."""
        
        if col_names is not None:
            self.columns = col_names
            return self
        
        header_idx = self.index[0] if header_idx is None else header_idx
        
        columns = self.v.loc[header_idx]
        index = self.loc[header_idx+1:].index
        
        return DDF(data=self.loc[header_idx+1:].values, index=index, columns=columns)

    def setIndex(self, keys):
        index_array = self.v[keys]
        cols_to_drop = [keys] if isinstance(keys, str) else keys
        return DDF(self.set_index(index_array).drop(columns=cols_to_drop))
    
    def setStyle(self, keys: list=[], subset:pd.IndexSlice=pd.IndexSlice[:,:]):
        """Return the styled values DF."""
        def get_style_kwargs(value_dict):
            return {k:v for k,v in value_dict.items() if k in keys}

        df_style = self.v.style.format(precision=2)
        for row_i in self.index:
            for col in self.columns:
                df_style.set_properties(subset=pd.IndexSlice[row_i,col], **get_style_kwargs(self.loc[row_i, col]))
        return df_style
 

# GOOGLE DOCS IO
def get_sheet_names(url: str, class_name: str="goog-inline-block docs-sheet-tab-caption") -> list[str]:
    """Find all Google Spreadsheet Sheet names

    Args:
        url (str): URL to the Google Spreadhseet
        class_name (str, optional): Tag's 'class' name to search for that contains sheet names. 
            Defaults to "goog-inline-block docs-sheet-tab-caption".

    Returns:
        list[str]: List of sheet names.
    """
    # create soup object
    soup = BS(requests.get(url).text, features="html.parser")
    
    # find all sheet raw tags as list
    sheet_tags = soup.find_all(name="div", attrs={"class": class_name})
    
    return [tag.contents[0] for tag in sheet_tags]

def get_sheet_ids(url: str, last_sheet_name: str) -> list[str]:
    """Return Gsheets ID's based on initial page url and last sheet name.

    Args:
        url (str): URL of the initial page.
        last_sheet_name (str): Name of the last sheet.

    Returns:
        list[str]: List of sheet ID-s as string.
    """

    # Send a GET request to the Google Sheets URL
    r = requests.get(url)

    # Create a BeautifulSoup object from the HTML content
    soup = BS(r.content, 'html.parser')
    
    # find <body> -> <script> -> sheet_names_pattern -> id_pattern
    pattern = re.compile(rf"topsnapshot.+{last_sheet_name}\\\"\]")
    sheet_ids = None
    for tag in soup.body.find_all('script'):
        
        if tag.string is not None:
            string = re.search(pattern, tag.text)
            if string is not None:
                sheet_ids = re.findall(r"\d{9,}", string[0])
                break
    
    # assert if first ID match matches the first sheet id from the url
    assert sheet_ids[0] == re.search(r"\d+?$", url)[0], "Gsheet ID's do not match with sheet names!"        
    
    return sheet_ids

def get_sheet_urls(url: str) -> dict:
    """Return dictionary of based on home url.

    Args:
        url (str): Home page URL.

    Returns:
        dict: Dict of {sheet_name:sheet_url}.
    """
    # match the base url w/o ending sheet ID
    base_url = re.match(r".*edit#gid=", url)[0]
    
    sheet_names = get_sheet_names(url)
    sheet_ids = get_sheet_ids(url, last_sheet_name=sheet_names[-1])
    
    # check that # of sheets equals # of found sheet IDs 
    assert len(sheet_names) == len(sheet_ids), "Number of sheets doesn't match with number of ID-s."
    
    # create a dictionary of sheet name : sheet_url
    sheet_urls = [base_url + sheet_id for sheet_id in sheet_ids]
    
    return {sheet_name: sheet_url for sheet_name,sheet_url in zip(sheet_names, sheet_urls)}

def read_gsheet(url:str, **read_csv_kwargs):
    """Read data from google sheets into DF."""
    
    # replace 'edit#gid' with 'export?format=csv&gid'
    url_csv = url.replace("edit#gid", "export?format=csv&gid")
    return pd.read_csv(url_csv, **read_csv_kwargs)

def downloadWorkbook(spreadsheet_url: str, 
                     file_path: str,
                     file_name: str='workbook') -> None:
    """Download Google Spreadsheet as an .xlsx file.
    """
    
    # Define the file name and extension
    file_ext = ".xlsx"

    # Generate a timestamp string in the format YYYY-MM-DD_HH-MM-SS
    timestamp = getTimestamp()

    # Concatenate the timestamp string with the file name and extension
    timestamped_file_name = f"{file_path}/{timestamp}_{file_name}{file_ext}"

    # SAVE GOOGLE SPREADSHEET AS .XLSX FILE
    export_url = spreadsheet_url.replace("edit#gid", "export?format=xlsx")

    response = requests.get(export_url)

    with open(timestamped_file_name, "wb") as output_file:
        output_file.write(response.content)
        
def downloadSheet(url: str, 
                  file_path: str,
                  route: str,
                  file_ext: str=".xlsx") -> None:
    """Download Google Spreadsheet as an .xlsx file.
    """

    # Generate a timestamp string in the format YYYY-MM-DD_HH-MM-SS
    timestamp = getTimestamp()
    full_file_name = f"{timestamp}_{route}{file_ext}"
    timestamped_file_name = os.path.join(file_path, full_file_name)
    
    # SAVE GOOGLE SPREADSHEET AS .XLSX FILE
    download_url = url.replace('/edit#','/export?format=xlsx&')

    response = requests.get(download_url)

    with open(timestamped_file_name, "wb") as output_file:
        output_file.write(response.content)

# LOCAL FILES IO
def getTimestamp(format: str="%Y-%m-%d_%H:%M:%S") -> str:
    """Generate a timestamp string in the format YYYY-MM-DD_HH:MM:SS""" 
    return datetime.datetime.now().strftime(format)

def getOldestFilename(files: list[str], ts_format:str="%Y-%m-%d_%H:%M:%S") -> str:
    """Returns the filename with oldest timestamp"""
    ts_pattern = r"\d[\d\-_:]+\d"
    return min(files, key=lambda x: datetime.datetime.strptime(re.search(ts_pattern, x)[0][:19], ts_format))

def getNewestFilename(files: list[str], ts_format:str="%Y-%m-%d_%H:%M:%S") -> str:
    """Returns the filename with newest timestamp"""
    ts_pattern = r"\d[\d\-_:]+\d"
    return max(files, key=lambda x: datetime.datetime.strptime(re.search(ts_pattern, x)[0][:19], ts_format))

def getPreviousFilename(files: list[str], ts_format:str="%Y-%m-%d_%H:%M:%S") -> str:
    """Returns the filename with oldest timestamp"""
    
    if len(files) == 1: 
        return files[0]
    
    return sorted(files, key=lambda x: datetime.datetime.strptime(x[:19], ts_format))[1] 

def logError(route:str, exception:Exception) -> None:
    """Save occurred error traceback to file with timestamp."""
    
    log_path = os.path.join("dashboard", "logs", "routes", route)
    if not os.path.exists(log_path):
        os.mkdir(log_path)
    
    files = os.listdir(log_path)
    while len(files) > 9:
        oldest_file = getOldestFilename(files=files)
        os.remove(os.path.join(log_path, oldest_file))
        
    error_name = exception.__class__.__name__
    
    with open(f'{log_path}/{getTimestamp()}_{error_name}', 'a') as f: 
        traceback.print_exc(file=f)


# .XLSX IO
def readXlsx(route: str, newest_file: bool=True) -> DDF:

    # Open the downloaded .xlsx file with openpyxl
    path_to_file = os.path.join("dashboard", "cache", route)
    file_name = getNewestFilename(os.listdir(path_to_file)) if newest_file else getPreviousFilename(os.listdir(path_to_file))
    full_rel_path = os.path.join(path_to_file, file_name)

    # read .xlsx w/ openpyxl & select the worksheet
    workbook = openpyxl.load_workbook(full_rel_path, data_only=True, read_only=True)
    sheet_name = workbook.sheetnames[0]
    worksheet = workbook[sheet_name]

    # read .xslx with pandas & get number of rows & cols accurately
    df_pandas = pd.read_excel(io=full_rel_path, sheet_name=sheet_name, header=None)
    max_row = df_pandas.index.max() + 1
    max_col = df_pandas.columns.max() + 1

    # Read in the rows
    rows = [row for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col)]

    row_values = [] # create an empty list to hold the cell values
    # iterate over the rows and columns to capture each cell properties
    for row in rows:
        cell_values = []
        for cell in row:
            # get the cell value, color, fill & font style
            cell_value = np.nan if cell.data_type in ['f'] or cell.value is None else cell.value
            
            fill_obj = cell.fill
            cell_color = fill_obj.start_color.index if not isinstance(fill_obj, NoneType) else np.nan
            cell_color = 'FFFFFFFF' if cell_color in ['00000000', None, NoneType, np.nan] else cell_color
            
            #cell_fill = cell.fill.fill_type if cell.fill else np.nan
            
            cell_font = cell.font
            
            font_weight = 'bold' if not isinstance(cell_font, NoneType) and cell_font.b else 'normal'
            
            # append the cell value, color, and fill to the list
            cell_values.append({'value': cell_value, 
                                'color': f"#{cell_color[2:]}" if isinstance(cell_color, str) else np.nan,
                                'font-weight': font_weight})

        row_values.append(cell_values)

    # close workbook when read_only=True
    if workbook.read_only:
        workbook.close()

    # create a pandas dataframe from the list of cell values
    return DDF(row_values)

### DATAFRAME IO ###
def findRefRowCol(df: pd.DataFrame, pattern: str, method: str='contains') -> tuple[int, int]:
    """Find reference row and column numeric index values.

    Args:
        pattern (str): Character sequence or regular expression.
        method (str, optional): Finding reference via pattern within the text ('contains')
            or equalling the value exactly ('equals'). Defaults to 'contains'.

    Raises:
        ValueError: Raises error if other than ['contains', 'equals'] is specified for the method.

    Returns:
        tuple[int, int]: Tuple of ['row_i', 'col_i']
    """
    
    
    # validate that method is correctly entered
    if method not in ['contains', 'equals']:
        raise ValueError(f"{method} can take only values: 'contains' or 'equals'!")
    
    idx_series = None
    if method == 'contains':
        idx_series = df.apply(lambda x: x.str.contains(pattern)).idxmax()
    if method == 'equals':
        idx_series = (df == pattern).idxmax()
    
    return idx_series.max(), idx_series.idxmax() 
    
def sliceDF(df: pd.DataFrame, 
            row_idx1: int=None, 
            row_idx2:int=None, 
            col_idx1: int=None, 
            col_idx2: int=None,
            col_0: bool=True, 
            direction='down') -> pd.DataFrame:
    """Slice DF till the first occuring empty row in given direction.

    Args:
        col_0 (bool, optional): If col_idx1 is actually first column. Defaults to True.
        direction (str, optional): Slice upwards or downwards from given row index. Defaults to 'down'.

    Raises:
        ValueError: If no row indices are specified.

    Returns:
        pd.DataFrame
    """
    
    # assert that at least one of the row indices is specified
    if row_idx1 is None and row_idx2 is None:
        raise ValueError(f"Both row indices can't equal {None}!") 
    
    # if column index is not the first column then None
    col_idx1 = col_idx1 if col_0 is True else None
    
    # find missing row index
    if direction == 'down':
        nan_mask = (df.iloc[row_idx1:,col_idx1:col_idx2] == 'nan').all(axis='columns')
        row_idx2 = None if nan_mask.sum() == 0 else nan_mask.idxmax()
        
    
    if direction == 'up':
        nan_mask = (df.iloc[:row_idx2,col_idx1:col_idx2] == 'nan').all(axis='columns')
        row_idx1 = None if nan_mask.sum() == 0 else nan_mask[::-1].idxmax() + 1
        row_idx2 += 1
         
    return df.iloc[row_idx1:row_idx2,col_idx1:col_idx2]

def getDFs(df: pd.DataFrame, references_dict: dict[str, tuple[str,str,str]]) -> dict[str, pd.DataFrame]:
    """Get subset DFs from a bigger DF based on reference strings.

    Args:
        df (pd.DataFrame): Raw initial DF.
        references_dict (dict): Dictionary of df_name: tuple('method', 'string' , 'direction', int('col_idx1)).
            method has 2 options ['contains' and 'equals'], 'direction' has 3 options 
            ['one', 'down', 'up'] where one means only one line needs to be parsed, 'up' and 'down'
            respectively correspond to the parsing direction from the reference point.
            
            Returns:
        dict: Dictionary of {'df_name': df}
    """
    
    
    # set all cols str type, NaN -> 'nan'
    df = df.astype(str)
    
    # find reference indices
    dfs = {}
    for k,v in references_dict.items():
        
        # find ref position
        row_i, col_i = findRefRowCol(df, pattern=v[1], method=v[0])
        
        # shift ref position if specified
        col_i = v[3] if len(v) == 4 else col_i
        
        if v[2] == 'one':
            dfs[k] = df.iloc[row_i,col_i:]
        if v[2] == 'down':
            dfs[k] = sliceDF(df, row_idx1=row_i, col_idx1=col_i, direction=v[2])
        if v[2] == 'up':
            dfs[k] = sliceDF(df, row_idx2=row_i, col_idx1=col_i, direction=v[2])

    # strip all NaN cols and rows
    dfs_clean = {}
    for k,df_ in dfs.items():
        
        if isinstance(df_, pd.Series):
            dfs_clean[k] = (df_
            .replace({'nan': np.nan})
            .dropna(how='any', axis='rows')
            )
        else:
            dfs_clean[k] = (df_
                .replace({'nan': np.nan})
                .dropna(how='all', axis='columns')
                .dropna(how='all', axis='rows')
            )
    
    return dfs_clean

def getNonSelectedDF(
    df_raw: pd.DataFrame, 
    df_dict: dict['df_name': pd.DataFrame], 
    start_idx: int=0, 
    stop_idx: int=None) -> pd.DataFrame:
    """Get data into DF that wasn't captured based on negative df_dict. 

    Args:
        df_raw (pd.DataFrame): Original DF where data wasn't cpatured.
        df_dict (_type_): Dict of the form {subdf_name: 'method', 'string' , 'direction', int('col_idx1))}.
        start_idx (int, optional): Starting row index (not positional).
        stop_idx (int, optional): Ending row index (not positional).

    Returns:
        pd.DataFrame: DF with rest of the info captured and NaN rows/cols stripped.
    """
    
    
    # find indicies already capture by "getDFs" function
    indices = np.array([])
    for df_ in df_dict.values():
        indices = np.concatenate((indices, df_.index.values))
    indices = [int(i) for i in indices]
    
    # subset DF
    df_sub = (df_raw
        .loc[~df_stocks_raw.index.isin(indices),:]
        .loc[start_idx:stop_idx,]
        .dropna(axis='columns', how='all')
        .dropna(axis='rows', how='all')
    )
    
    return df_sub

def getRelContainsIdx(df: pd.DataFrame, pattern: str, na=None):
    """Return relative row index where pattern is found."""
    return np.where(df.apply(lambda x: x.str.contains(pattern, na=np.nan) == True))[0][0]

### DF MODIFICATION ###
def total_assets(df: pd.DataFrame, index:str, col:str) -> pd.DataFrame:
    """Calculate total value if '#ERROR!'
    """
 
    # check if total value has ERROR msg
    if df.loc[index, col] in ["#ERROR!", "#NAME?"]:
        df.loc[index, col] = 0
    else:
        return df.loc[index, col]
     
    # exclude error fields and monthly income
    df = df[(~df[col].isin(['#ERROR!', '#NAME?'])) & (df.index != 'Monthly Income')]
    
    # str to float and calc total sum
    df.loc[:, col] = df[col].replace(r"[\$,]", "", regex=True).astype(float)
    df.loc[index, col] = df[col].sum()
    
    # format numbers back to string
    return "${0:,.2f}".format(df.loc[index, col])

def comment_button(series:pd.Series) -> str:
    """Replaces value in the DF to given html.

    Args:
        sereis (pd.Series): pandas Series obj
        data : data

    Returns:
        str: Returns the html string
    """
    return f'<button type="button" class="btn btn-secondary btn-sm" data-bs-content="{series} '\
        '"data-bs-toggle="popover" data-bs-trigger="focus" '\
        'style="--bs-btn-font-size: .85rem;">Details</button>'

def total_value_to_num(df: pd.DataFrame):
    """Modify DF data column to numeric
    """
    
    # add underscores to col names
    df.columns = df.columns.map(lambda x: x.replace(" ", "_"))
    
    # set 'Asset_Class' as new index
    df = (df
        .set_index('Asset_Class') 
        .drop('Comments', axis='columns')
        .loc[:'Total (USD)'].iloc[:-1]
        .query("Total_Value != '#ERROR!'")
    )
    
    df['Total_Value'] = pd.to_numeric(df['Total_Value'].replace(r"[\$,]", "", regex=True),
                                      errors='coerce')
    return df.reset_index()

def formatToDollars(value, precision: int=2):
    """Format number to have commas separating thousans and leading $ sign."""
    if value == np.nan or pd.isna(value):
        return value
    try:
        float_value = float(value)
        return f"${float_value:,.{precision}f}"
    except ValueError:
        return value

def addButton(df: pd.DataFrame,
                  col_names: list[str]=None,
                  popover_style_str: str="",
                  button_name: str="Details",
                  button_class: str="btn btn-secondary btn-sm",
                  data_bs_html: str="true",
                  data_bs_toggle: str="popover",
                  data_bs_trigger: str="focus",
                  button_style: str="--bs-btn-font-size: .85rem;"):
        """Add button html to the value-dict."""
        def add_button(value):
            if value in NONE_LIKE_LIST:
                return value
            
            html_string = " ".join(
                (
                f'<button type="button" class="{button_class}"', 
                f'data-bs-content="<div style="{popover_style_str}">{value}</div>"',
                f'data-bs-html="{data_bs_html}" data-bs-toggle={data_bs_toggle} data-bs-trigger="{data_bs_trigger}"', 
                f'style="{button_style}">{button_name}</button>'
                )
            )
            return html_string 
        
        # apply to specified columns
        df[col_names] = df[col_names].applymap(add_button)
        
        # return entire DF
        return df
    
# Adds colors column dynamically based on risk assessment
def riskPallette(series: pd.Series, scale: dict) -> pd.Series:
    """Apply color based on risk level (# between 0-100).

    Args:
        scale (dict): Dictionary of the form {0:color, ... , 9:color}

    Returns:
        pd.Series: color
    """
    if series >= 90: 
        return scale["9"]
    if series < 10: 
        return scale["0"]
    else: 
        return scale[str(series)[0]]
    
### DF STYLING ###
def set_bg_color(val, cmap: dict) -> str:
    """Map colors to DF values based on mapping dictionary.

    Args:
        val ( any type): Any value type.
        cmap (dict): Dict of the form {val:'color'}
    """
    return f'background-color: {cmap[val]}'

def styleDf(df: pd.DataFrame, route: str):
    """Style DF using colors in .xlsx file.

    Args:
        df (pd.DataFrame): Must contain original row indices.
        route (str): Name of the route in cache folder, e.g. 'stocks_watchlist'

    """
    ## ORIGINAL DF ## 
    i1, i2 = df.columns.name if not None else df.index[0]-1, df.index[-1] + 1
    
    ## EXCEL TABLE ##
    path_to_file = f'dashboard/cache/{route}/'
    newest_file = getNewestFilename(os.listdir(path_to_file))
    full_path = path_to_file + newest_file
    
    # load workbook
    wb = openpyxl.load_workbook(filename=full_path, data_only=True)
    ws = wb.active

    # define the range of cells to read, use original df references
    rows = ws.iter_rows(max_row=i2)

    # create an empty list to hold the cell values
    row_values = []

    # iterate over the rows and columns to capture each cell properties
    for row in rows:
        cell_values = []
        for cell in row:
            # get the cell value, color, and fill
            cell_value = "" if cell.data_type in ['f'] or cell.value is None else cell.value
            cell_color = cell.fill.start_color.index if cell.fill.start_color else np.nan
            cell_fill = cell.fill.fill_type if cell.fill else np.nan

            # append the cell value, color, and fill to the list
            cell_values.append({'value': cell_value, 
                                'color': f"#{cell_color[2:]}" if isinstance(cell_color, str) else np.nan, 
                                'fill': cell_fill})

        row_values.append(cell_values)
    
    # create a pandas dataframe from the list of cell values
    df_style_dict = pd.DataFrame(row_values)
    
    new_idx = np.insert(df.index.values, 0, i1)
    df_style_dict = df_style_dict.loc[new_idx,] # get original rows
    
    df_style_dict_cols = df_style_dict.iloc[0,].apply(lambda x: x['value']) # get all cols
    df_style_dict = df_style_dict.iloc[1:,]
    df_style_dict.columns = df_style_dict_cols
    df_style_dict = df_style_dict.loc[:, df.columns] # get original rows & cols
    
    # check if respective DFs match
    assert df.shape == df_style_dict.shape, "Provided DF and Excel Table have different shape."
    assert df.index.equals(df_style_dict.index), "Provided DF and Excel Table have different row indices."
    assert df.columns.equals(df_style_dict.columns), "Provided DF and Excel Table have different column labels."
    
    # style original DF based on Excel
    styled_df = df.fillna("").style
    for i, row in df_style_dict.iterrows():
        for c, value in row.items():
            styled_dct = {'color': '#FFFFFF'} if value["color"] in ['#000000', None, np.nan] \
                else {'color': f'{value["color"]}'}
            styled_df.set_properties(subset=pd.IndexSlice[i, c], **styled_dct)
    
    return styled_df

### HELPERS ###
def get_risk_pallete(pallette: dict) -> dict['int':'color']:
    """Generate risk pallete in scale 0 to 100 in steps of 10.
    Args:
        pallette (dict): Dictionary {n: ['colors'....]}
    Returns:
        Dict: Dictionary {"0": 'color'}
    """
    return {str(i):color for i,color in enumerate(pallette(10)[::-1])}

# DECORATORS
def ioCacheAndLog(
    url: str,
    route: str,
    testing: bool=False):
    
    def my_decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            
            if not testing:
            
                # delete oldest file if 3 files present
                cache_path = os.path.join("dashboard", "cache", route)
                if not os.path.exists(cache_path):
                    os.mkdir(cache_path)
                
                files = os.listdir(cache_path)
                while len(files) > 2:
                    oldest_file = getOldestFilename(files=files)
                    files = [file for file in files if file != oldest_file]
                    os.remove(os.path.join(cache_path, oldest_file))
                
                # download google sheet as .xlsx file
                downloadSheet(url=url, file_path=cache_path, route=route)
            
            try:
                ddf = readXlsx(route=route, newest_file=True)
                result = func(ddf)
                # if any errors occured except block is executed
                return result
            
            # read data from cache
            except Exception as e:
                
                if not testing:
                    # capture and save occurred error log
                    logError(route=route, exception=e)
                else:
                    print(e)
                
                # read data from the previous cached file
                ddf = readXlsx(route=route, newest_file=False)
                newest_file = getNewestFilename(files)
                os.remove(os.path.join(cache_path, newest_file))
                
                result = func(ddf)
                return result
        return wrapper
    
    return my_decorator
   
### ROUTE SPECIFIC ###
# INVESTMENTS #
def calcTotalUSD(df: pd.DataFrame, col_name:str) -> pd.DataFrame:
    """Calculate total value if '#ERROR!' in col_name."""
    total_idx = np.where(df.apply(lambda x: x.str.contains(r"Total \(USD\)") == True))[0][0]
    exclude_cols = ['Monthly Income', np.nan]
 
    # check if total value has ERROR msg
    if df.loc[total_idx, col_name] in ["#ERROR!", "#NAME?"]:
        df.loc[total_idx, col_name] = 0
    else:
        return df
     
    # exclude error fields and monthly income
    df_mask = df[(~df[col_name].isin(['#ERROR!', '#NAME?'])) & (~df['Asset Class'].isin(exclude_cols))]
    
    # str to float and calc total sum
    df_mask.loc[col_name] = df_mask[col_name].replace(r"[\$,]", "", regex=True).astype(float)
    
    total = df_mask[col_name].sum()
    df.loc[total_idx, col_name] = "${0:,.2f}".format(total) # format numbers back to string

    return df

